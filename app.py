from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import json
import os
from datetime import datetime, timedelta
import requests
import tempfile
import subprocess
from werkzeug.utils import secure_filename
import pandas as pd
import ffmpeg
import logging
import io
import vlc
import sys
import platform
import openpyxl
from openpyxl.styles import PatternFill
import csv
from video_utils import convert_wmv_to_mp4, upload_to_gcs
from google.cloud import storage
from google.cloud import firestore
from dotenv import load_dotenv
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.cell_range import CellRange

# Load environment variables
load_dotenv()

app = Flask(__name__)
CORS(app)
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500MB max file size
app.config['UPLOAD_FOLDER'] = 'uploads'

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize VLC instance with proper error handling
try:
    # Create a basic vlc instance
    vlc_instance = vlc.Instance()
    if not vlc_instance:
        raise Exception("Failed to create VLC instance")
        
    # Create an empty vlc media player
    player = vlc_instance.media_player_new()
    if not player:
        raise Exception("Failed to create media player")
        
    logger.info("VLC initialized successfully")
except Exception as e:
    logger.error(f"Error initializing VLC: {str(e)}")
    vlc_instance = None
    player = None

# Store usage statistics
USAGE_STATS = {
    'BI': 0,
    'BV': 0,
    'VI': 0,
    'VV': 0,
    'SRC': 0
}

# Ensure the uploads directory exists
UPLOADS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
if not os.path.exists(UPLOADS_DIR):
    os.makedirs(UPLOADS_DIR)

AUDD_API_TOKEN = '04e48a84490a8a2f0bf327b274404905'  # Replace with your actual API token

# Initialize Firestore client with proper error handling
try:
    # Use default credentials (works on Google Cloud if service account has permissions)
    firestore_client = firestore.Client()
    MUSIC_CO_COLLECTION = 'music_companies'
    # Test the connection
    test_doc = firestore_client.collection(MUSIC_CO_COLLECTION).limit(1).get()
    logger.info("Firestore initialized successfully")
except Exception as e:
    logger.error(f"Error initializing Firestore: {str(e)}")
    firestore_client = None

@app.route('/')
def index():
    return render_template('splash.html')

@app.route('/main')
def main():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'video' not in request.files:
            return jsonify({'error': 'No video file provided'}), 400
        
        file = request.files['video']
        if file.filename == '':
            return jsonify({'error': 'No selected file'}), 400

        # Check file size
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        
        if file_size > app.config['MAX_CONTENT_LENGTH']:
            return jsonify({
                'error': f'File too large. Maximum size is {app.config["MAX_CONTENT_LENGTH"] / (1024*1024)}MB'
            }), 413

        # Create uploads directory if it doesn't exist
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        
        # Save the uploaded file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{secure_filename(file.filename)}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        # Check if it's a WMV file
        if filename.lower().endswith('.wmv'):
            try:
                # Convert WMV to MP4
                output_filename = f"{os.path.splitext(filename)[0]}.mp4"
                output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
                
                # Convert the file
                convert_wmv_to_mp4(filepath, output_path)
                
                # Upload to GCS
                gcs_filename = f"videos/{output_filename}"
                upload_to_gcs(output_path, gcs_filename)
                
                # Clean up temporary files
                os.remove(filepath)
                os.remove(output_path)
                
                # Generate signed URL
                url = generate_signed_url(gcs_filename)
                return jsonify({'filename': url})
                
            except Exception as e:
                # Clean up on error
                if os.path.exists(filepath):
                    os.remove(filepath)
                if os.path.exists(output_path):
                    os.remove(output_path)
                return jsonify({'error': f'Error converting video: {str(e)}'}), 500
        else:
            # For non-WMV files, just upload to GCS
            try:
                gcs_filename = f"videos/{filename}"
                upload_to_gcs(filepath, gcs_filename)
                
                # Clean up temporary file
                os.remove(filepath)
                
                # Generate signed URL
                url = generate_signed_url(gcs_filename)
                return jsonify({'filename': url})
                
            except Exception as e:
                # Clean up on error
                if os.path.exists(filepath):
                    os.remove(filepath)
                return jsonify({'error': f'Error uploading video: {str(e)}'}), 500
                
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/uploads/<filename>')
def serve_video(filename):
    return send_file(os.path.join(UPLOADS_DIR, filename))

@app.route('/api/markers', methods=['GET', 'POST'])
def handle_markers():
    if request.method == 'POST':
        markers = request.json
        # Save markers to a file
        with open('markers.json', 'w') as f:
            json.dump(markers, f)
        return jsonify({'status': 'success'})
    else:
        # Load markers from file
        try:
            with open('markers.json', 'r') as f:
                markers = json.load(f)
            return jsonify(markers)
        except FileNotFoundError:
            return jsonify([])

@app.route('/api/export/<format>', methods=['POST'])
def export_markers(format):
    try:
        payload = request.json
        header_rows = payload.get('headerRows', [])
        markers = payload.get('markers', [])
        blank_lines = payload.get('blankLines', 0)
        fields_to_export = payload.get('fieldsToExport')
        field_labels = payload.get('fieldLabels', {})
        time_format = payload.get('timeFormat', 'HH:MM:SS')  # Default to HH:MM:SS
        metadata_with_format = payload.get('metadata_with_format', None)
        table_header_format = payload.get('table_header_format', None)
        table_row_format = payload.get('table_row_format', None)
        column_widths = payload.get('column_widths', None)

        def format_time(seconds, include_frames=False):
            hours = int(seconds // 3600)
            minutes = int((seconds % 3600) // 60)
            secs = int(seconds % 60)
            if include_frames:
                frames = int(round((seconds % 1) * 25))  # Assuming 25 fps
                return f"{hours:02d}:{minutes:02d}:{secs:02d}:{frames:02d}"
            return f"{hours:02d}:{minutes:02d}:{secs:02d}"

        def convert_time_fields(row):
            new_row = row.copy()
            def is_timecode_string(val):
                import re
                return isinstance(val, str) and re.match(r"^\d{2}:\d{2}:\d{2}(:\d{2})?$", val)

            # Handle TCR In
            if 'tcrIn' in new_row and new_row['tcrIn']:
                if is_timecode_string(new_row['tcrIn']):
                    if time_format == 'HH:MM:SS' and len(new_row['tcrIn'].split(':')) == 4:
                        new_row['tcrIn'] = ':'.join(new_row['tcrIn'].split(':')[:3])
                else:
                    try:
                        new_row['tcrIn'] = format_time(float(new_row['tcrIn']), time_format == 'HH:MM:SS:FF')
                    except (ValueError, TypeError):
                        new_row['tcrIn'] = ''

            # Handle TCR Out
            if 'tcrOut' in new_row and new_row['tcrOut']:
                if is_timecode_string(new_row['tcrOut']):
                    if time_format == 'HH:MM:SS' and len(new_row['tcrOut'].split(':')) == 4:
                        new_row['tcrOut'] = ':'.join(new_row['tcrOut'].split(':')[:3])
                else:
                    try:
                        new_row['tcrOut'] = format_time(float(new_row['tcrOut']), time_format == 'HH:MM:SS:FF')
                    except (ValueError, TypeError):
                        new_row['tcrOut'] = ''

            # Handle Duration
            if 'duration' in new_row and new_row['duration']:
                if is_timecode_string(new_row['duration']):
                    if time_format == 'HH:MM:SS' and len(new_row['duration'].split(':')) == 4:
                        new_row['duration'] = ':'.join(new_row['duration'].split(':')[:3])
                else:
                    try:
                        new_row['duration'] = format_time(float(new_row['duration']), time_format == 'HH:MM:SS:FF')
                    except (ValueError, TypeError):
                        new_row['duration'] = ''

            return new_row

        if format == 'excel':
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_path = temp_file.name
            temp_file.close()

            wb = openpyxl.Workbook()
            ws = wb.active

            # --- BEGIN: Write metadata rows with formatting ---
            if metadata_with_format:
                merge_ranges = set()
                for row_idx, row in enumerate(metadata_with_format, 1):
                    for col_idx, cell in enumerate(row, 1):
                        value = cell.get('value', '')
                        bold = cell.get('bold', False)
                        align = cell.get('align', None)
                        merge = cell.get('merge', None)
                        ws.cell(row=row_idx, column=col_idx, value=value)
                        if bold:
                            ws.cell(row=row_idx, column=col_idx).font = Font(bold=True)
                        if align:
                            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal=align)
                        if merge and merge not in merge_ranges:
                            ws.merge_cells(merge)
                            merge_ranges.add(merge)
            else:
                for row in header_rows:
                    ws.append(row)
            # --- END: Write metadata rows ---

            # Add blank lines if specified
            for _ in range(blank_lines):
                ws.append([])

            # --- BEGIN: Write marker table with formatting ---
            table_start_row = ws.max_row + 1
            if markers:
                if fields_to_export:
                    if 'seq' not in fields_to_export:
                        fields_to_export.insert(0, 'seq')
                    if 'seq' not in field_labels:
                        field_labels['seq'] = 'SEQ#'
                    # Write header row
                    ws.append([field_labels.get(field, field) for field in fields_to_export])
                    # Apply header formatting
                    if table_header_format:
                        for col_idx, fmt in enumerate(table_header_format, 1):
                            cell = ws.cell(row=table_start_row, column=col_idx)
                            if fmt.get('bold'): cell.font = Font(bold=True)
                            if fmt.get('align'): cell.alignment = Alignment(horizontal=fmt.get('align'))
                            if fmt.get('fill') and fmt.get('fill') != '00000000':
                                cell.fill = PatternFill(start_color=fmt.get('fill'), end_color=fmt.get('fill'), fill_type='solid')
                            if fmt.get('border') and 'openpyxl' in fmt.get('border'):
                                # Simple border: apply thin border to all sides
                                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    # Write data rows
                    for i, row in enumerate(markers):
                        row = convert_time_fields(row)
                        row_data = [i + 1]
                        if 'usage' in row:
                            if isinstance(row['usage'], list):
                                row['usage'] = ','.join(row['usage'])
                            elif isinstance(row['usage'], str) and row['usage'].startswith('[') and row['usage'].endswith(']'):
                                try:
                                    arr = json.loads(row['usage'])
                                    if isinstance(arr, list):
                                        row['usage'] = ','.join(arr)
                                except Exception:
                                    pass
                        row_data.extend([row.get(field, '') for field in fields_to_export[1:]])
                        ws.append(row_data)
                        # Apply data row formatting
                        if table_row_format:
                            for col_idx, fmt in enumerate(table_row_format, 1):
                                cell = ws.cell(row=table_start_row + 1 + i, column=col_idx)
                                if fmt.get('bold'): cell.font = Font(bold=True)
                                if fmt.get('align'): cell.alignment = Alignment(horizontal=fmt.get('align'))
                                if fmt.get('fill') and fmt.get('fill') != '00000000':
                                    cell.fill = PatternFill(start_color=fmt.get('fill'), end_color=fmt.get('fill'), fill_type='solid')
                                if fmt.get('border') and 'openpyxl' in fmt.get('border'):
                                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                else:
                    keys = ['seq'] + list(markers[0].keys())
                    ws.append(keys)
                    for i, row in enumerate(markers):
                        row = convert_time_fields(row)
                        row_data = [i + 1]
                        if 'usage' in row:
                            if isinstance(row['usage'], list):
                                row['usage'] = ','.join(row['usage'])
                            elif isinstance(row['usage'], str) and row['usage'].startswith('[') and row['usage'].endswith(']'):
                                try:
                                    arr = json.loads(row['usage'])
                                    if isinstance(arr, list):
                                        row['usage'] = ','.join(arr)
                                except Exception:
                                    pass
                        row_data.extend(list(row.values()))
                        ws.append(row_data)
                        # Apply data row formatting
                        if table_row_format:
                            for col_idx, fmt in enumerate(table_row_format, 1):
                                cell = ws.cell(row=table_start_row + 1 + i, column=col_idx)
                                if fmt.get('bold'): cell.font = Font(bold=True)
                                if fmt.get('align'): cell.alignment = Alignment(horizontal=fmt.get('align'))
                                if fmt.get('fill') and fmt.get('fill') != '00000000':
                                    cell.fill = PatternFill(start_color=fmt.get('fill'), end_color=fmt.get('fill'), fill_type='solid')
                                if fmt.get('border') and 'openpyxl' in fmt.get('border'):
                                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            # --- END: Write marker table with formatting ---

            # Set column widths
            if column_widths:
                for idx, width in enumerate(column_widths, 1):
                    if width:
                        ws.column_dimensions[get_column_letter(idx)].width = width

            # --- BEGIN: Color marked rows ---
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            start_row = len(header_rows) + blank_lines + 2 if markers else len(header_rows) + blank_lines + 1
            for i, marker in enumerate(markers):
                color = marker.get('markColor', '')
                if color == 'yellow':
                    for cell in ws[start_row + i]:
                        cell.fill = yellow_fill
                elif color == 'red':
                    for cell in ws[start_row + i]:
                        cell.fill = red_fill
            # --- END: Color marked rows ---

            wb.save(temp_path)
            with open(temp_path, 'rb') as f:
                file_data = f.read()

            # Use original file name for export if provided
            export_file_name = payload.get('originalFileName', 'exported_file')
            return send_file(
                io.BytesIO(file_data),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'{export_file_name}.xlsx'
            )

        elif format == 'csv':
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.csv')
            temp_path = temp_file.name
            temp_file.close()

            with open(temp_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                # --- BEGIN: Write metadata rows with formatting ---
                if metadata_with_format:
                    merge_ranges = set()
                    for row_idx, row in enumerate(metadata_with_format, 1):
                        for col_idx, cell in enumerate(row, 1):
                            value = cell.get('value', '')
                            bold = cell.get('bold', False)
                            align = cell.get('align', None)
                            merge = cell.get('merge', None)
                            writer.writerow([value])
                            if bold:
                                writer.writerow([value]).font = Font(bold=True)
                            if align:
                                writer.writerow([value]).alignment = Alignment(horizontal=align)
                            if merge and merge not in merge_ranges:
                                writer.writerow([value]).merge = merge
                                merge_ranges.add(merge)
                else:
                    for row in header_rows:
                        writer.writerow(row)
                # --- END: Write metadata rows ---

                # Add blank lines if specified
                for _ in range(blank_lines):
                    writer.writerow([])

                # --- BEGIN: Write marker table ---
                if markers:
                    if fields_to_export:
                        # Add SEQ# to the beginning of fields_to_export if not already present
                        if 'seq' not in fields_to_export:
                            fields_to_export.insert(0, 'seq')
                        
                        # Add SEQ# to field labels if not present
                        if 'seq' not in field_labels:
                            field_labels['seq'] = 'SEQ#'
                        
                        writer.writerow([field_labels.get(field, field) for field in fields_to_export])
                        for i, row in enumerate(markers):
                            # Convert time fields to the selected format
                            row = convert_time_fields(row)
                            # Add SEQ# to the beginning of the row data
                            row_data = [i + 1]  # SEQ# is index + 1
                            # Convert usage array or stringified array to comma-separated string if it exists
                            if 'usage' in row:
                                if isinstance(row['usage'], list):
                                    row['usage'] = ','.join(row['usage'])
                                elif isinstance(row['usage'], str) and row['usage'].startswith('[') and row['usage'].endswith(']'):
                                    try:
                                        arr = json.loads(row['usage'])
                                        if isinstance(arr, list):
                                            row['usage'] = ','.join(arr)
                                    except Exception:
                                        pass
                            # Add the rest of the fields
                            row_data.extend([row.get(field, '') for field in fields_to_export[1:]])
                            writer.writerow(row_data)
                    else:
                        # Add SEQ# to the beginning of the keys
                        keys = ['seq'] + list(markers[0].keys())
                        writer.writerow(keys)
                        for i, row in enumerate(markers):
                            # Convert time fields to the selected format
                            row = convert_time_fields(row)
                            # Add SEQ# to the beginning of the row data
                            row_data = [i + 1]  # SEQ# is index + 1
                            # Convert usage array or stringified array to comma-separated string if it exists
                            if 'usage' in row:
                                if isinstance(row['usage'], list):
                                    row['usage'] = ','.join(row['usage'])
                                elif isinstance(row['usage'], str) and row['usage'].startswith('[') and row['usage'].endswith(']'):
                                    try:
                                        arr = json.loads(row['usage'])
                                        if isinstance(arr, list):
                                            row['usage'] = ','.join(arr)
                                    except Exception:
                                        pass
                            # Add the rest of the values
                            row_data.extend(list(row.values()))
                            writer.writerow(row_data)
                # --- END: Write marker table ---

            # Use original file name for export if provided
            export_file_name = payload.get('originalFileName', 'exported_file')
            return send_file(
                io.BytesIO(open(temp_path, 'rb').read()),
                mimetype='text/csv',
                as_attachment=True,
                download_name=f'{export_file_name}.csv'
            )

        else:
            return jsonify({'error': 'Invalid export format'}), 400

    except Exception as e:
        logger.error(f"Export failed: {e}")
        return jsonify({'error': f'Export failed: {str(e)}'}), 500
    finally:
        if 'temp_path' in locals() and os.path.exists(temp_path):
            try:
                os.unlink(temp_path)
            except Exception as cleanup_error:
                logger.error(f"Error cleaning up temporary file: {str(cleanup_error)}")

@app.route('/api/usage-stats', methods=['GET', 'POST'])
def handle_usage_stats():
    global USAGE_STATS
    if request.method == 'POST':
        usage = request.json.get('usage')
        if usage in USAGE_STATS:
            USAGE_STATS[usage] += 1
    return jsonify(USAGE_STATS)

@app.route('/api/recognize-audio', methods=['POST'])
def recognize_audio():
    try:
        tcr_in = request.form.get('tcrIn')
        tcr_out = request.form.get('tcrOut')
        video_src = request.form.get('videoSrc')
        file = request.files.get('file')

        logger.info(f"Received recognize request - TCR In: {tcr_in}, TCR Out: {tcr_out}")
        logger.info(f"Video source: {video_src}, File: {file.filename if file else 'None'}")

        if not tcr_in or not tcr_out:
            return jsonify({'status': 'error', 'message': 'Missing TCR In or Out'}), 400

        # Calculate start and duration in seconds
        def time_to_seconds(t):
            h, m, s = map(float, t.split(':'))
            return int(h) * 3600 + int(m) * 60 + s
        
        start = time_to_seconds(tcr_in)
        end = time_to_seconds(tcr_out)
        duration = end - start
        
        if duration <= 0:
            return jsonify({'status': 'error', 'message': 'Invalid time range'}), 400

        # Prepare input file path
        if file:
            # Save uploaded file to temp
            try:
                # First, read the file into memory to check its size
                file_content = file.read()
                file_size = len(file_content)
                logger.info(f"Uploaded file size: {file_size} bytes")
                
                if file_size == 0:
                    return jsonify({
                        'status': 'error',
                        'message': 'The uploaded file is empty'
                    }), 400
                
                # Reset file pointer for saving
                file.seek(0)
                
                # Create a unique temporary directory for this upload
                upload_dir = tempfile.mkdtemp()
                input_path = os.path.join(upload_dir, 'input.mp4')
                logger.info(f"Created temporary directory: {upload_dir}")
                
                # Write the file in chunks to handle large files
                chunk_size = 8192
                bytes_written = 0
                with open(input_path, 'wb') as temp_in:
                    while True:
                        chunk = file.read(chunk_size)
                        if not chunk:
                            break
                        temp_in.write(chunk)
                        bytes_written += len(chunk)
                        logger.info(f"Upload progress: {bytes_written}/{file_size} bytes ({(bytes_written/file_size)*100:.2f}%)")
                
                logger.info(f"File upload completed. Total bytes written: {bytes_written}")
                
                if bytes_written != file_size:
                    logger.error(f"File size mismatch. Expected {file_size} bytes, got {bytes_written} bytes")
                    return jsonify({
                        'status': 'error',
                        'message': 'File upload was incomplete. Please try again.'
                    }), 400
                
                # Validate the video file
                try:
                    # First check if file exists and has content
                    if not os.path.exists(input_path) or os.path.getsize(input_path) == 0:
                        return jsonify({
                            'status': 'error',
                            'message': 'The video file is empty or incomplete. Please try uploading again.'
                        }), 400
                    
                    # Use ffprobe to get detailed video information
                    probe_cmd = [
                        'ffprobe',
                        '-v', 'error',
                        '-print_format', 'json',
                        '-show_format',
                        '-show_streams',
                        input_path
                    ]
                    
                    logger.info(f"Running ffprobe validation: {' '.join(probe_cmd)}")
                    probe_result = subprocess.run(probe_cmd, capture_output=True, text=True)
                    
                    if probe_result.returncode != 0:
                        logger.error(f"Invalid video file: {probe_result.stderr}")
                        return jsonify({
                            'status': 'error',
                            'message': 'The video file appears to be invalid or corrupted. Please try uploading again.'
                        }), 400
                    
                    # Log the probe result for debugging
                    logger.info(f"FFprobe result: {probe_result.stdout}")
                    
                    # Try to fix the video file using a different approach
                    try:
                        logger.info("Attempting to fix video file format")
                        fixed_path = os.path.join(upload_dir, 'fixed.mp4')
                        
                        # First try: Use -movflags +faststart
                        fix_cmd1 = [
                            'ffmpeg',
                            '-y',
                            '-v', 'error',
                            '-i', input_path,
                            '-c', 'copy',
                            '-movflags', '+faststart',
                            fixed_path
                        ]
                        
                        logger.info(f"Running first fix attempt: {' '.join(fix_cmd1)}")
                        fix_result1 = subprocess.run(fix_cmd1, capture_output=True, text=True)
                        
                        if fix_result1.returncode != 0:
                            logger.warning(f"First fix attempt failed: {fix_result1.stderr}")
                            
                            # Second try: Re-encode the video
                            logger.info("Attempting re-encoding fix")
                            fix_cmd2 = [
                                'ffmpeg',
                                '-y',
                                '-v', 'error',
                                '-i', input_path,
                                '-c:v', 'libx264',
                                '-c:a', 'aac',
                                '-movflags', '+faststart',
                                fixed_path
                            ]
                            
                            logger.info(f"Running second fix attempt: {' '.join(fix_cmd2)}")
                            fix_result2 = subprocess.run(fix_cmd2, capture_output=True, text=True)
                            
                            if fix_result2.returncode != 0:
                                logger.error(f"Second fix attempt failed: {fix_result2.stderr}")
                                return jsonify({
                                    'status': 'error',
                                    'message': 'Could not process the video file. Please try uploading a different file.'
                                }), 400
                        
                        # Replace original input path with fixed file
                        os.remove(input_path)
                        input_path = fixed_path
                        logger.info(f"Successfully fixed video file: {input_path}")
                        
                    except Exception as e:
                        logger.error(f"Error fixing video file: {str(e)}")
                        return jsonify({
                            'status': 'error',
                            'message': 'Error processing video file. Please try uploading again.'
                        }), 500
                    
                except Exception as e:
                    logger.error(f"Error validating video file: {str(e)}")
                    return jsonify({
                        'status': 'error',
                        'message': 'Error validating video file. Please try uploading again.'
                    }), 400
                    
            except Exception as e:
                logger.error(f"Error handling file upload: {str(e)}")
                return jsonify({
                    'status': 'error',
                    'message': f'Error handling file upload: {str(e)}'
                }), 500
        elif video_src:
            # Download the video file
            try:
                logger.info(f"Starting video download from URL: {video_src}")
                
                # Create a unique temporary directory for this download
                upload_dir = tempfile.mkdtemp()
                input_path = os.path.join(upload_dir, 'input.mp4')
                logger.info(f"Created temporary directory: {upload_dir}")
                
                # Get file size from headers if available
                head_response = requests.head(video_src, allow_redirects=True)
                content_length = head_response.headers.get('content-length')
                if content_length:
                    logger.info(f"Expected file size: {content_length} bytes")
                
                # Download the file with progress tracking
                response = requests.get(video_src, stream=True)
                response.raise_for_status()
                
                total_size = int(response.headers.get('content-length', 0))
                block_size = 8192
                downloaded_size = 0
                
                with open(input_path, 'wb') as temp_in:
                    for chunk in response.iter_content(chunk_size=block_size):
                        if chunk:
                            temp_in.write(chunk)
                            downloaded_size += len(chunk)
                            if total_size > 0:
                                progress = (downloaded_size / total_size) * 100
                                logger.info(f"Download progress: {downloaded_size}/{total_size} bytes ({progress:.2f}%)")
                
                logger.info(f"Download completed. Total bytes downloaded: {downloaded_size}")
                
                # Verify download is complete
                if content_length and int(content_length) != downloaded_size:
                    logger.error(f"Download incomplete. Expected {content_length} bytes, got {downloaded_size} bytes")
                    return jsonify({
                        'status': 'error',
                        'message': 'Video download was incomplete. Please try again.'
                    }), 400
                
                # Validate the downloaded file
                try:
                    # Check if file exists and has content
                    if not os.path.exists(input_path) or os.path.getsize(input_path) == 0:
                        return jsonify({
                            'status': 'error',
                            'message': 'The downloaded video file is empty or incomplete.'
                        }), 400
                    
                    # Use ffprobe to validate the video
                    probe_cmd = [
                        'ffprobe',
                        '-v', 'error',
                        '-print_format', 'json',
                        '-show_format',
                        '-show_streams',
                        input_path
                    ]
                    
                    logger.info(f"Running ffprobe validation: {' '.join(probe_cmd)}")
                    probe_result = subprocess.run(probe_cmd, capture_output=True, text=True)
                    
                    if probe_result.returncode != 0:
                        logger.error(f"Invalid video file: {probe_result.stderr}")
                        return jsonify({
                            'status': 'error',
                            'message': 'The downloaded video file appears to be invalid or corrupted.'
                        }), 400
                    
                    logger.info(f"FFprobe result: {probe_result.stdout}")
                    
                    # Try to fix the video file
                    try:
                        logger.info("Attempting to fix video file format")
                        fixed_path = os.path.join(upload_dir, 'fixed.mp4')
                        
                        # First try: Use -movflags +faststart
                        fix_cmd1 = [
                            'ffmpeg',
                            '-y',
                            '-v', 'error',
                            '-i', input_path,
                            '-c', 'copy',
                            '-movflags', '+faststart',
                            fixed_path
                        ]
                        
                        logger.info(f"Running first fix attempt: {' '.join(fix_cmd1)}")
                        fix_result1 = subprocess.run(fix_cmd1, capture_output=True, text=True)
                        
                        if fix_result1.returncode != 0:
                            logger.warning(f"First fix attempt failed: {fix_result1.stderr}")
                            
                            # Second try: Re-encode the video
                            logger.info("Attempting re-encoding fix")
                            fix_cmd2 = [
                                'ffmpeg',
                                '-y',
                                '-v', 'error',
                                '-i', input_path,
                                '-c:v', 'libx264',
                                '-c:a', 'aac',
                                '-movflags', '+faststart',
                                fixed_path
                            ]
                            
                            logger.info(f"Running second fix attempt: {' '.join(fix_cmd2)}")
                            fix_result2 = subprocess.run(fix_cmd2, capture_output=True, text=True)
                            
                            if fix_result2.returncode != 0:
                                logger.error(f"Second fix attempt failed: {fix_result2.stderr}")
                                return jsonify({
                                    'status': 'error',
                                    'message': 'Could not process the downloaded video file.'
                                }), 400
                        
                        # Replace original input path with fixed file
                        os.remove(input_path)
                        input_path = fixed_path
                        logger.info(f"Successfully fixed video file: {input_path}")
                        
                    except Exception as e:
                        logger.error(f"Error fixing video file: {str(e)}")
                        return jsonify({
                            'status': 'error',
                            'message': 'Error processing downloaded video file.'
                        }), 500
                    
                except Exception as e:
                    logger.error(f"Error validating downloaded file: {str(e)}")
                    return jsonify({
                        'status': 'error',
                        'message': 'Error validating downloaded video file.'
                    }), 400
                    
            except requests.RequestException as e:
                logger.error(f"Error downloading video: {str(e)}")
                return jsonify({
                    'status': 'error',
                    'message': f'Error downloading video: {str(e)}'
                }), 500
        else:
            return jsonify({
                'status': 'error',
                'message': 'No file or videoSrc provided'
            }), 400

        # Extract segment and convert to mp3
        with tempfile.NamedTemporaryFile(delete=False, suffix='.mp3') as temp_out:
            output_path = temp_out.name
            logger.info(f"Created output file: {output_path}")

        try:
            # Use ffmpeg-python to extract audio
            logger.info(f"Extracting audio segment from {start}s to {end}s")
            
            # Process the audio with more detailed logging
            logger.info("Starting FFmpeg processing...")
            
            # Use direct FFmpeg command with explicit options
            ffmpeg_cmd = [
                'ffmpeg',
                '-y',  # Overwrite output file
                '-v', 'error',  # Only show errors
                '-i', input_path,  # Input file
                '-ss', str(start),  # Start time
                '-t', str(duration),  # Duration
                '-vn',  # No video
                '-acodec', 'libmp3lame',  # Audio codec
                '-ab', '128k',  # Audio bitrate
                '-ac', '1',  # Mono audio
                '-ar', '44100',  # Sample rate
                '-f', 'mp3',  # Force MP3 format
                '-movflags', '+faststart',  # Move metadata to start of file
                output_path  # Output file
            ]
            
            logger.info(f"Running FFmpeg command: {' '.join(ffmpeg_cmd)}")
            
            # Run FFmpeg
            result = subprocess.run(ffmpeg_cmd, capture_output=True, text=True)
            
            if result.stderr:
                logger.warning(f"FFmpeg stderr output: {result.stderr}")
                # Check for specific error conditions
                if "moov atom not found" in result.stderr:
                    logger.error("Input video file appears to be incomplete or corrupted")
                    return jsonify({
                        'status': 'error',
                        'message': 'The video file appears to be incomplete or corrupted. Please ensure the video is fully uploaded before processing.'
                    }), 400
                elif "Invalid data found" in result.stderr:
                    logger.error("Invalid video data detected")
                    return jsonify({
                        'status': 'error',
                        'message': 'The video file contains invalid data. Please check the file format and try again.'
                    }), 400
            
            if result.returncode != 0:
                logger.error(f"FFmpeg failed with return code {result.returncode}")
                logger.error(f"FFmpeg error output: {result.stderr}")
                raise Exception(f"FFmpeg failed with return code {result.returncode}")
            
            # Verify the output file exists and has content
            if not os.path.exists(output_path):
                raise Exception("FFmpeg output file was not created")
            
            file_size = os.path.getsize(output_path)
            logger.info(f"FFmpeg output file size: {file_size} bytes")
            
            if file_size == 0:
                raise Exception("FFmpeg output file is empty")
            
            logger.info("Audio extraction completed successfully")
            
        except subprocess.CalledProcessError as e:
            logger.error(f"FFmpeg error: {e.stderr}")
            return jsonify({'status': 'error', 'message': 'ffmpeg error', 'stderr': e.stderr}), 500
        except Exception as e:
            logger.error(f"Error processing audio: {str(e)}")
            return jsonify({'status': 'error', 'message': f'Error processing audio: {str(e)}'}), 500

        # Send to audd.io with enhanced error handling
        try:
            logger.info("Sending audio to audd.io")
            with open(output_path, 'rb') as f:
                files = {'file': f}
                data = {
                    'api_token': AUDD_API_TOKEN,
                    'return': 'apple_music,spotify'
                }
                r = requests.post('https://api.audd.io/', data=data, files=files)
                r.raise_for_status()
                result = r.json()
                
                # Log the complete response for debugging
                logger.info(f"Complete audd.io response: {json.dumps(result, indent=2)}")
                
                if result.get('status') == 'error':
                    error_message = result.get('error', {}).get('message', 'Unknown error from audd.io')
                    logger.error(f"audd.io API error: {error_message}")
                    return jsonify({
                        'status': 'error',
                        'message': f'audd.io API error: {error_message}',
                        'details': result
                    }), 500
                
                logger.info(f"Received response from audd.io: {result.get('status', 'unknown')}")
                
        except requests.RequestException as e:
            logger.error(f"Error calling audd.io: {str(e)}")
            result = {'status': 'error', 'message': f'Error calling audd.io: {str(e)}'}
        except Exception as e:
            logger.error(f"Error processing audd.io response: {str(e)}")
            result = {'status': 'error', 'message': 'Invalid response from audd.io'}

        # Clean up temp files
        try:
            os.remove(output_path)
            if file or video_src:
                # Remove the entire temporary directory
                if 'upload_dir' in locals():
                    import shutil
                    shutil.rmtree(upload_dir)
            logger.info("Cleaned up temporary files")
        except Exception as e:
            logger.error(f"Error cleaning up files: {str(e)}")

        return jsonify(result)

    except Exception as e:
        logger.error(f"Unexpected error in recognize_audio: {str(e)}")
        return jsonify({'status': 'error', 'message': f'Unexpected error: {str(e)}'}), 500

@app.route('/api/parse-cue-sheet', methods=['POST'])
def parse_cue_sheet():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        file = request.files['file']
        filename = secure_filename(file.filename)
        ext = filename.split('.')[-1].lower()
        metadata = []
        metadata_with_format = []
        header = []
        data = []
        table_header_format = []
        table_row_format = []
        column_widths = []
        if ext in ['xlsx', 'xls']:
            import openpyxl
            from openpyxl.utils import get_column_letter, range_boundaries
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.worksheet.cell_range import CellRange
            import io
            wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
            ws = wb.active
            # Helper to get merged cell value
            def get_merged_cell_value(ws, cell):
                for merged_range in ws.merged_cells.ranges:
                    min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                    if min_row <= cell.row <= max_row and min_col <= cell.column <= max_col:
                        return ws.cell(row=min_row, column=min_col).value
                return cell.value
            # Get merged cell ranges
            merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=6), 1):
                row_data = []
                row_format = []
                for cell in row:
                    value = get_merged_cell_value(ws, cell)
                    cell_info = {
                        'value': value,
                        'bold': cell.font.bold if cell.font else False,
                        'align': cell.alignment.horizontal if cell.alignment else None,
                        'merge': None
                    }
                    for rng in merged_ranges:
                        cr = CellRange(rng)
                        if (cell.row, cell.column) in cr.cells:
                            cell_info['merge'] = rng
                            break
                    row_data.append(value if value is not None else '')
                    row_format.append(cell_info)
                metadata.append(row_data)
                metadata_with_format.append(row_format)
            # Table header (row 7)
            try:
                table_header_row = list(ws.iter_rows(min_row=7, max_row=7))[0]
                for cell in table_header_row:
                    value = get_merged_cell_value(ws, cell)
                    cell_info = {
                        'bold': cell.font.bold if cell.font else False,
                        'align': cell.alignment.horizontal if cell.alignment else None,
                        'fill': cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None,
                        'border': str(cell.border) if cell.border else None
                    }
                    table_header_format.append(cell_info)
            except Exception as e:
                table_header_format = []
            # First data row (row 8)
            try:
                table_data_row = list(ws.iter_rows(min_row=8, max_row=8))[0]
                for cell in table_data_row:
                    value = get_merged_cell_value(ws, cell)
                    cell_info = {
                        'bold': cell.font.bold if cell.font else False,
                        'align': cell.alignment.horizontal if cell.alignment else None,
                        'fill': cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None,
                        'border': str(cell.border) if cell.border else None
                    }
                    table_row_format.append(cell_info)
            except Exception as e:
                table_row_format = []
            # Column widths
            try:
                for col in ws.columns:
                    col_letter = get_column_letter(col[0].column)
                    width = ws.column_dimensions[col_letter].width
                    column_widths.append(width)
            except Exception as e:
                column_widths = []
            # Get header and data as before, but fill merged cell values
            rows = list(ws.iter_rows(values_only=False))
            # Filter out rows where all cells are empty
            rows = [row for row in rows if any(str(get_merged_cell_value(ws, cell)).strip() for cell in row)]
            header = [str(get_merged_cell_value(ws, cell)) if get_merged_cell_value(ws, cell) is not None else '' for cell in rows[6]] if len(rows) > 6 else []
            data_rows = rows[7:] if len(rows) > 7 else []
            data = [dict(zip(header, [str(get_merged_cell_value(ws, cell)) if get_merged_cell_value(ws, cell) is not None else '' for cell in row])) for row in data_rows]
        elif ext == 'csv':
            import pandas as pd
            df = pd.read_csv(file, header=None)
            df = df.fillna('')
            rows = df.values.tolist()
            # Filter out rows where all cells are empty
            rows = [row for row in rows if any(str(cell).strip() for cell in row)]
            metadata = [[str(cell) if cell is not None else '' for cell in row] for row in rows[:6]]
            metadata_with_format = [[{'value': str(cell) if cell is not None else '', 'bold': False, 'align': None, 'merge': None} for cell in row] for row in rows[:6]]
            header = [str(cell) if cell is not None else '' for cell in rows[6]] if len(rows) > 6 else []
            data_rows = rows[7:] if len(rows) > 7 else []
            data = [dict(zip(header, [str(cell) if cell is not None else '' for cell in row])) for row in data_rows]
            table_header_format = [{'bold': False, 'align': None, 'fill': None, 'border': None} for _ in header]
            table_row_format = [{'bold': False, 'align': None, 'fill': None, 'border': None} for _ in header]
            column_widths = [None for _ in header]
        else:
            return jsonify({'error': 'Unsupported file type'}), 400
        return jsonify({
            'metadata': metadata,
            'metadata_with_format': metadata_with_format,
            'header': header,
            'data': data,
            'table_header_format': table_header_format,
            'table_row_format': table_row_format,
            'column_widths': column_widths
        })
    except Exception as e:
        logging.exception("Failed to parse cue sheet file")
        return jsonify({'error': str(e)}), 400

@app.route('/test-ffmpeg')
def test_ffmpeg():
    try:
        result = subprocess.run(['ffmpeg', '-version'], capture_output=True, text=True)
        if result.returncode == 0:
            return jsonify({
                'status': 'success',
                'message': 'FFmpeg is installed',
                'version': result.stdout.split('\n')[0]
            })
        else:
            return jsonify({
                'status': 'error',
                'message': 'FFmpeg command failed',
                'error': result.stderr
            }), 500
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Error checking FFmpeg: {str(e)}'
        }), 500

@app.route('/api/vlc/play', methods=['POST'])
def vlc_play():
    try:
        video_path = request.json.get('video_path')
        if not video_path:
            return jsonify({'error': 'No video path provided'}), 400

        # Create media from path
        media = vlc_instance.media_new(video_path)
        player.set_media(media)
        player.play()
        
        return jsonify({
            'status': 'success',
            'message': 'Video started playing',
            'duration': player.get_length() / 1000  # Convert to seconds
        })
    except Exception as e:
        logger.error(f"Error playing video with VLC: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/vlc/pause', methods=['POST'])
def vlc_pause():
    try:
        player.pause()
        return jsonify({'status': 'success', 'message': 'Video paused'})
    except Exception as e:
        logger.error(f"Error pausing video: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/vlc/stop', methods=['POST'])
def vlc_stop():
    try:
        player.stop()
        return jsonify({'status': 'success', 'message': 'Video stopped'})
    except Exception as e:
        logger.error(f"Error stopping video: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/vlc/seek', methods=['POST'])
def vlc_seek():
    try:
        position = request.json.get('position')  # Position in seconds
        if position is None:
            return jsonify({'error': 'No position provided'}), 400
            
        # Convert position to milliseconds
        position_ms = int(position * 1000)
        player.set_time(position_ms)
        
        return jsonify({'status': 'success', 'message': f'Seeked to {position} seconds'})
    except Exception as e:
        logger.error(f"Error seeking video: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/vlc/status', methods=['GET'])
def vlc_status():
    try:
        state = player.get_state()
        position = player.get_time() / 1000  # Convert to seconds
        duration = player.get_length() / 1000  # Convert to seconds
        
        return jsonify({
            'status': 'success',
            'state': str(state),
            'position': position,
            'duration': duration
        })
    except Exception as e:
        logger.error(f"Error getting VLC status: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/vlc/volume', methods=['POST'])
def vlc_volume():
    try:
        volume = request.json.get('volume')
        if volume is None:
            return jsonify({'error': 'No volume provided'}), 400
            
        # Set volume (0-100)
        player.audio_set_volume(int(volume * 100))
        
        return jsonify({'status': 'success', 'message': f'Volume set to {volume * 100}%'})
    except Exception as e:
        logger.error(f"Error setting volume: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/export-settings')
def export_settings():
    return render_template('export_settings.html')

@app.route('/generate-upload-url', methods=['POST'])
def generate_upload_url():
    data = request.json
    filename = data['filename']
    bucket_name = 'mos-aat'
    storage_client = storage.Client()
    bucket = storage_client.bucket(bucket_name)
    blob = bucket.blob(f'uploads/{filename}')
    url = blob.generate_signed_url(
        version='v4',
        expiration=timedelta(minutes=15),
        method='PUT',
        content_type='application/octet-stream',
    )
    return jsonify({'url': url})

@app.route('/process-uploaded-video', methods=['POST'])
def process_uploaded_video():
    data = request.json
    filename = data['filename']
    bucket_name = 'mos-aat'
    storage_client = storage.Client()
    bucket = storage_client.bucket(bucket_name)
    blob = bucket.blob(f'uploads/{filename}')
    local_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    blob.download_to_filename(local_path)
    # Convert if WMV, else just copy
    if filename.lower().endswith('.wmv'):
        output_filename = f"{os.path.splitext(filename)[0]}.mp4"
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
        convert_wmv_to_mp4(local_path, output_path)
        # Upload converted file
        converted_blob = bucket.blob(f'converted/{output_filename}')
        converted_blob.upload_from_filename(output_path)
        # Clean up
        os.remove(local_path)
        os.remove(output_path)
        # Generate signed URL for converted file
        url = converted_blob.generate_signed_url(
            version='v4',
            expiration=timedelta(hours=1),
            method='GET',
        )
        return jsonify({'converted_url': url})
    else:
        # For non-WMV, just move to converted/
        converted_blob = bucket.blob(f'converted/{filename}')
        converted_blob.rewrite(blob)
        url = converted_blob.generate_signed_url(
            version='v4',
            expiration=timedelta(hours=1),
            method='GET',
        )
        return jsonify({'converted_url': url})

@app.route('/api/music-co', methods=['GET'])
def list_music_co():
    search = request.args.get('search', '').lower()
    docs = firestore_client.collection(MUSIC_CO_COLLECTION).stream()
    results = []
    for doc in docs:
        data = doc.to_dict()
        name = data.get('name', '')
        if not search or search in name.lower():
            results.append({'id': doc.id, 'name': name})
    return jsonify(results)

@app.route('/api/music-co', methods=['POST'])
def add_music_co():
    data = request.json
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Name is required'}), 400
    # Prevent duplicates
    docs = firestore_client.collection(MUSIC_CO_COLLECTION).where('name', '==', name).stream()
    if any(True for _ in docs):
        return jsonify({'error': 'Music Co already exists'}), 400
    doc_ref = firestore_client.collection(MUSIC_CO_COLLECTION).document()
    doc_ref.set({'name': name})
    return jsonify({'id': doc_ref.id, 'name': name})

@app.route('/api/music-co/<id>', methods=['DELETE'])
def delete_music_co(id):
    doc_ref = firestore_client.collection(MUSIC_CO_COLLECTION).document(id)
    if not doc_ref.get().exists:
        return jsonify({'error': 'Music Co not found'}), 404
    doc_ref.delete()
    return jsonify({'success': True})

@app.route('/music-co')
def music_co_manager():
    return render_template('music_co.html')

if __name__ == '__main__':
    app.run(debug=True) 